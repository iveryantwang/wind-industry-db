# -*- coding: utf-8 -*-
"""
风电行业数据库 —— Excel 导出脚本
=================================
读取 vault 内所有笔记的 YAML frontmatter，导出为与「储能行业数据库」同版式的 xlsx。

用法：
    python export_excel.py                    # 导出到 vault 根目录
    python export_excel.py <vault路径>
    python export_excel.py <vault路径> <输出文件>

依赖：openpyxl
"""

import os
import sys
import io
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 复用 rebuild_tables 的 frontmatter 解析
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from rebuild_tables import read_notes  # noqa: E402

# ---------------------------------------------------------------- 配色（沿用附件版式）

C_DARK = "142640"      # 标题栏深底
C_SUB = "A0B4C8"       # 副标题字
C_HEAD = "1B3A5C"      # 表头
C_HEAD2 = "2E5077"     # 次级表头
C_TXT = "1A1A2E"       # 正文
C_MUTED = "5A5A6E"     # 弱化文字
C_ZEBRA = "F0F4F8"     # 斑马纹
C_LINK = "1B3A6B"

KPI_COLORS = [
    ("累计条目", "1B3A5C"),
    ("客户与项目", "1B3A6B"),
    ("竞争对手", "8B6914"),
    ("风电政策", "2D5016"),
    ("技术方案", "7A1F3D"),
    ("投资财务", "3D2A5C"),
    ("每日新闻", "2E5077"),
    ("更新日期", "3D6EA5"),
]

REGIONS = [
    ("欧洲", "1B3A6B", "DCE7F3"),
    ("亚太", "8B6914", "F5E8D0"),
    ("北美", "2D5016", "DDEBD0"),
    ("拉美", "7A1F3D", "F0DCE0"),
    ("中东非", "6B4423", "F0E6D2"),
    ("全球", "3D2A5C", "E6E0F0"),
]

PRIORITY_FILL = {"P0": ("C62828", "FFFFFF"), "P1": ("E65100", "FFFFFF"),
                 "P2": ("F9A825", C_TXT), "P3": ("2E7D32", "FFFFFF")}

IMPORTANCE_FILL = {"高": ("C62828", "FFFFFF"), "中": ("F9A825", C_TXT), "低": ("90A4AE", "FFFFFF")}

# 机会等级：高=可争取的大单，已失单=整机已被对手拿下
OPPORTUNITY_FILL = {"高": ("2E7D32", "FFFFFF"), "中": ("F9A825", C_TXT),
                    "低": ("90A4AE", "FFFFFF"), "已失单": ("757575", "FFFFFF")}

FONT = "微软雅黑"
THIN = Side(style="thin", color="D8DEE6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------- 表定义

SHEETS = [
    {
        "name": "1-每日新闻", "type": "news", "sort": ("date", True),
        "cols": [
            ("date", "日期", 12), ("title", "标题", 42), ("summary", "摘要", 55),
            ("source_name", "来源", 16), ("region", "区域", 10), ("country", "国别", 12),
            ("segment", "板块", 10), ("importance", "重要性", 9), ("source_url", "原文链接", 40),
        ],
    },
    {
        "name": "2-客户与项目", "type": "project", "sort": ("capacity_mw", True),
        "cols": [
            ("opportunity", "机会等级", 10), ("region", "区域", 10), ("country", "国别", 12),
            ("customer", "业主/客户", 30), ("project", "项目名称", 32),
            ("description", "项目简介", 50),
            ("capacity_mw", "装机容量(MW)", 13), ("unit_mw", "单机容量(MW)", 13),
            ("turbine_count", "台数", 8), ("turbine_model", "机型", 18),
            ("segment", "类型", 10), ("status", "项目状态", 11),
            ("oem", "整机商(空=未定)", 18), ("cod_year", "并网年", 9),
            ("capex", "投资规模", 18),
            ("owner_contact", "我方对接人", 13), ("source_url", "来源", 36),
        ],
    },
    {
        "name": "3-风电政策", "type": "policy", "sort": ("issued", True),
        "cols": [
            ("region", "区域", 10), ("country", "国别", 12), ("policy", "政策名称", 32),
            ("issued", "发布日期", 12), ("issuer", "发布机构", 24),
            ("summary", "核心内容", 50), ("impact", "影响评估", 38),
            ("impact_level", "影响等级", 10), ("source_url", "来源", 36),
        ],
    },
    {
        "name": "4-竞争对手", "type": "competitor", "sort": ("date", True),
        "cols": [
            ("region", "区域", 10), ("country", "国别", 12), ("competitor", "竞争对手", 22),
            ("event_type", "动态类型", 12), ("counterparty", "签约客户", 25),
            ("contract_value", "合同金额", 16), ("capacity_mw", "容量(MW)", 11),
            ("turbine_count", "台数", 8), ("date", "日期", 12), ("source_url", "来源", 36),
        ],
    },
    {
        "name": "5-技术方案", "type": "tech", "sort": ("date", True),
        "cols": [
            ("region", "区域", 10), ("country", "国别", 12), ("solution", "技术方案/场景", 28),
            ("description", "描述", 50), ("use_case", "应用场景", 22),
            ("provider", "提供方", 22), ("maturity", "成熟度", 10),
            ("date", "日期", 12), ("source_url", "来源", 36),
        ],
    },
    {
        "name": "6-投资财务", "type": "finance", "sort": ("date", True),
        "cols": [
            ("region", "区域", 10), ("country", "国别", 12), ("subject", "项目/模型", 30),
            ("capex", "投资规模", 20), ("irr", "IRR/回报率", 14), ("lcoe", "LCOE", 20),
            ("biz_model", "商业模式", 35), ("funding", "资金来源", 40),
            ("date", "日期", 12), ("source_url", "来源", 36),
        ],
    },
    {
        "name": "7-战略判断", "type": "strategy", "sort": ("priority", False),
        "cols": [
            ("priority", "优先级", 10), ("region", "区域", 10), ("country", "国别", 12),
            ("owner_party", "业主/开发商", 30), ("project", "项目", 32),
            ("capacity_mw", "容量(MW)", 11), ("action", "建议行动", 55),
            ("rationale", "判断依据", 55),
            ("deadline", "行动窗口", 12), ("status", "状态", 11),
        ],
    },
]

NUMERIC = {"capacity_mw", "unit_mw", "turbine_count", "cod_year"}


# ---------------------------------------------------------------- 工具

def val(v):
    if v is None:
        return ""
    if isinstance(v, list):
        return " / ".join(str(x) for x in v)
    return str(v).strip()


def num_or_text(field, raw):
    s = val(raw)
    if not s:
        return None
    if field in NUMERIC:
        try:
            f = float(s)
            return int(f) if f == int(f) else f
        except ValueError:
            return s
    return s


def sort_key(field):
    def k(n):
        v = n.get(field)
        if v in (None, "", []):
            return (1, 0, "")
        try:
            return (0, float(str(v)), "")
        except (ValueError, TypeError):
            return (0, 0, str(v))
    return k


def style_header(ws, headers, widths):
    for i, (h, wd) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------- 数据表

def build_sheet(wb, spec, notes):
    ws = wb.create_sheet(spec["name"])
    headers = [h for (_, h, _) in spec["cols"]]
    widths = [w for (_, _, w) in spec["cols"]]
    style_header(ws, headers, widths)

    field, desc = spec["sort"]
    rows = sorted(notes, key=sort_key(field), reverse=desc)
    if desc:
        filled = [n for n in rows if n.get(field) not in (None, "", [])]
        empty = [n for n in rows if n.get(field) in (None, "", [])]
        rows = filled + empty

    for r, n in enumerate(rows, start=2):
        zebra = (r % 2 == 1)
        for i, (fld, _h, _w) in enumerate(spec["cols"], start=1):
            v = num_or_text(fld, n.get(fld))
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=9, color=C_TXT)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            if zebra:
                c.fill = PatternFill("solid", fgColor=C_ZEBRA)

            if fld in ("region", "country", "date", "issued", "deadline", "first_logged",
                       "segment", "status", "maturity", "event_type", "impact_level",
                       "importance", "priority", "cod_year", "turbine_count", "opportunity"):
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if fld in ("capacity_mw", "unit_mw"):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0.#"

            # 链接列
            if fld == "source_url" and isinstance(v, str) and v.startswith("http"):
                c.hyperlink = v
                c.font = Font(name=FONT, size=8, color="1B6BC0", underline="single")

            # 优先级 / 重要性 色块
            if fld == "priority" and val(v) in PRIORITY_FILL:
                bg, fg = PRIORITY_FILL[val(v)]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name=FONT, size=9, bold=True, color=fg)
            if fld in ("importance", "impact_level") and val(v) in IMPORTANCE_FILL:
                bg, fg = IMPORTANCE_FILL[val(v)]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name=FONT, size=9, bold=True, color=fg)
            if fld == "opportunity" and val(v) in OPPORTUNITY_FILL:
                bg, fg = OPPORTUNITY_FILL[val(v)]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name=FONT, size=9, bold=True, color=fg)
            # 整机商留空 = 尚未选定，高亮提示
            if fld == "oem" and not val(v) and val(n.get("opportunity")) in ("高", "中"):
                c.value = "未定 ★"
                c.fill = PatternFill("solid", fgColor="FFF3C4")
                c.font = Font(name=FONT, size=9, bold=True, color="8B6914")
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 30

    last = max(len(rows) + 1, 2)
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), last)
    return len(rows)


# ---------------------------------------------------------------- 总览

def build_overview(wb, counts, today):
    ws = wb.create_sheet("总览", 0)
    for i, w_ in enumerate([16, 12, 10, 45, 12, 10, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w_

    def merge_set(rng, value, size, bold, fg, bg, align="center"):
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = value
        c.font = Font(name=FONT, size=size, bold=bold, color=fg)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return c

    # 标题区
    merge_set("A1:H1", "全球风电行业数据库", 20, True, "FFFFFF", C_DARK)
    merge_set("A2:H2",
              "Global Wind Power Industry Intelligence Database  |  更新时间：%s" % today,
              11, False, C_SUB, C_DARK)
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 12

    # KPI 卡片：两行，每行 4 张，公式引用各表
    kpi_formulas = [
        "=COUNTA('1-每日新闻'!A2:A5000)+COUNTA('2-客户与项目'!A2:A5000)"
        "+COUNTA('3-风电政策'!A2:A5000)+COUNTA('4-竞争对手'!A2:A5000)"
        "+COUNTA('5-技术方案'!A2:A5000)+COUNTA('6-投资财务'!A2:A5000)"
        "+COUNTA('7-战略判断'!A2:A5000)",
        "=COUNTA('2-客户与项目'!A2:A5000)",
        "=COUNTA('4-竞争对手'!A2:A5000)",
        "=COUNTA('3-风电政策'!A2:A5000)",
        "=COUNTA('5-技术方案'!A2:A5000)",
        "=COUNTA('6-投资财务'!A2:A5000)",
        "=COUNTA('1-每日新闻'!A2:A5000)",
        today,
    ]
    pos = [("A4:B4", "A5:B5"), ("C4:D4", "C5:D5"), ("E4:F4", "E5:F5"), ("G4:H4", "G5:H5"),
           ("A7:B7", "A8:B8"), ("C7:D7", "C8:D8"), ("E7:F7", "E8:F8"), ("G7:H7", "G8:H8")]
    for (label, color), (lr, vr), f in zip(KPI_COLORS, pos, kpi_formulas):
        merge_set(lr, label, 9, True, "FFFFFF", color)
        merge_set(vr, f, 18 if not str(f).startswith("2026") else 14, True, "FFFFFF", color)
    for r, h in [(4, 22), (5, 36), (6, 8), (7, 22), (8, 36), (9, 12)]:
        ws.row_dimensions[r].height = h

    # 区域分布（基于客户与项目）
    merge_set("A10:H10", "区域分布概览（客户与项目）", 12, True, "FFFFFF", C_HEAD, "left")
    ws.row_dimensions[10].height = 28
    for i, h in enumerate(["区域", "项目数", "占比", "数据条"], start=1):
        c = ws.cell(row=11, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD2)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 24

    total_row = 11 + len(REGIONS) + 1
    for j, (name, fc, bg) in enumerate(REGIONS):
        r = 12 + j
        a = ws.cell(row=r, column=1, value=name)
        a.font = Font(name=FONT, size=10, bold=True, color=fc)
        a.fill = PatternFill("solid", fgColor=bg)
        a.alignment = Alignment(horizontal="center", vertical="center")

        # 注意：客户与项目表 A 列是「机会等级」，B 列才是「区域」，此处必须指向 B 列
        b = ws.cell(row=r, column=2, value='=COUNTIF(\'2-客户与项目\'!B:B,A%d)' % r)
        b.font = Font(name=FONT, size=10, bold=True, color=C_TXT)
        b.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=r, column=3, value="=IF($B$%d=0,0,B%d/$B$%d)" % (total_row, r, total_row))
        c.number_format = "0.0%"
        c.font = Font(name=FONT, size=10, color=C_TXT)
        c.alignment = Alignment(horizontal="center", vertical="center")

        d = ws.cell(row=r, column=4,
                    value='=IF(B%d=0,"-",REPT("█",MAX(1,ROUND(B%d/MAX($B$%d,1)*32,0))))'
                          % (r, r, total_row))
        d.font = Font(name=FONT, size=10, color=fc)
        d.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22

    for i, v in enumerate(["合计",
                           "=SUM(B12:B%d)" % (total_row - 1),
                           "=IF($B$%d=0,0,1)" % total_row, ""], start=1):
        c = ws.cell(row=total_row, column=i, value=v)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if i == 3:
            c.number_format = "0%"
    ws.row_dimensions[total_row].height = 24

    # 板块明细
    sec_head = total_row + 2
    merge_set("A%d:D%d" % (sec_head, sec_head), "板块数据明细", 12, True, "FFFFFF", C_HEAD, "left")
    ws.row_dimensions[sec_head].height = 28
    for i, h in enumerate(["板块", "条目数", "占比", "更新方式"], start=1):
        c = ws.cell(row=sec_head + 1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD2)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sec_head + 1].height = 24

    sections = [
        ("一、每日风电新闻", "1-每日新闻", "🤖 每日自动更新"),
        ("二、我们的客户与项目", "2-客户与项目", "🤖 每日自动更新"),
        ("三、各国风电政策", "3-风电政策", "🤖 每日自动更新"),
        ("四、竞争对手情况", "4-竞争对手", "🤖 每日自动更新"),
        ("五、技术方案情况", "5-技术方案", "🤖 每日自动更新"),
        ("六、投资财务情况", "6-投资财务", "🤖 每日自动更新"),
        ("七、战略判断", "7-战略判断", "🤖 每日自动更新"),
    ]
    first = sec_head + 2
    for j, (label, sheet, how) in enumerate(sections):
        r = first + j
        bg = "FFFFFF" if j % 2 == 0 else C_ZEBRA
        cells = [
            (1, label, Font(name=FONT, size=10, bold=True, color=C_TXT), "left"),
            (2, "=COUNTA('%s'!A2:A5000)" % sheet,
             Font(name=FONT, size=10, bold=True, color="3D6EA5"), "center"),
            (3, "=IF($A$5=0,0,B%d/$A$5)" % r,
             Font(name=FONT, size=10, color=C_MUTED), "center"),
            (4, how, Font(name=FONT, size=9, color=C_MUTED), "left"),
        ]
        for col, v, ft, al in cells:
            c = ws.cell(row=r, column=col, value=v)
            c.font = ft
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=al, vertical="center")
            if col == 3:
                c.number_format = "0.0%"
        ws.row_dimensions[r].height = 22

    # 脚注
    foot = first + len(sections) + 1
    merge_set("A%d:H%d" % (foot, foot),
              "数据来源：WebSearch 多引擎检索 + 行业媒体定向  |  区域结构：欧洲 / 亚太 / 北美 / 拉美 / 中东非 / 全球  |  "
              "自动化：Claude 每日全板块自动更新（含客户与项目、战略判断）；抓取范围为前一日信息  |  "
              "本表由 99-系统/scripts/export_excel.py 从 Obsidian 笔记自动生成，请勿直接编辑本文件",
              8, False, C_MUTED, C_ZEBRA, "left")
    ws.row_dimensions[foot].height = 34
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------- 主流程

def main():
    here = os.path.dirname(os.path.abspath(__file__))
    vault = os.path.abspath(sys.argv[1]) if len(sys.argv) > 1 else os.path.abspath(os.path.join(here, "..", ".."))
    today = datetime.date.today().isoformat()
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(vault, "风电行业数据库_%s.xlsx" % today)

    notes = read_notes(vault)
    by_type = {}
    for n in notes:
        by_type.setdefault(n["type"], []).append(n)

    wb = Workbook()
    wb.remove(wb.active)

    counts = {}
    for spec in SHEETS:
        counts[spec["name"]] = build_sheet(wb, spec, by_type.get(spec["type"], []))

    build_overview(wb, counts, today)
    wb.active = 0

    try:
        wb.save(out)
    except PermissionError:
        # 目标文件被 Excel 打开时会锁定，退回到带时间戳的备用文件名
        base, ext = os.path.splitext(out)
        stamp = datetime.datetime.now().strftime("%H%M%S")
        out = "%s_%s%s" % (base, stamp, ext)
        wb.save(out)
        print("! 目标文件被占用（可能正在 Excel 中打开），已改存为备用文件名")
    print("已导出: %s" % out)
    for k, v in counts.items():
        print("  %-14s %d 条" % (k, v))
    print("  合计 %d 条" % sum(counts.values()))
    return out


if __name__ == "__main__":
    main()
